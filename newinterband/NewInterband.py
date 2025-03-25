from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import ast
import re, os
import time


# Function to extract the desired data from the header file
def extract_labels_from_header(header_file):
    with open(header_file, 'r') as file:
        content = file.read()
    # Parse the content as a Python literal (it will safely convert the string into a list)
    data = ast.literal_eval(content)

    # Extract only the second elements from each inner list (e.g., 'R1', 'Y1', etc.)
    extracted_labels = [item[1] for item in data]

    # Return the extracted labels in the desired format
    return extracted_labels

def custom_sort(lst):
    order={'R':0,'Y':1,'B':2}
    return sorted(lst,key=lambda x:(order[x[0]],x[1:]))


# Function to create Excel sheet
def create_excel_sheet(filename, header_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "2487_1"
    ws.merge_cells("C1:O1")
    ws["C1"] = "KRE1012487_1 Interband Isolation Sample-3"
    ws["C1"].font = Font(bold=True)
    ws["C1"].alignment = Alignment(horizontal="center")
    ws["C1"].fill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")

    # Use the new function to extract labels
    extracted_label = extract_labels_from_header(header_file)
    print(f"from text retrival{extracted_label}")
    res=custom_sort(extracted_label)
    print(f"from custom sort{res}")

    a, b = [], []
    for i in res:
        a.append(i)
        a.append("")
        b.append("P45")
        b.append("M45")
    headers = ["PORT", ""] + a
    sub_headers = ["", ""] + b
    ws.append(headers)
    print(headers)
    ws.append(sub_headers)
    print(sub_headers)

    for col in range(3, len(headers) + 1):
        ws.cell(row=2, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=3, column=col).alignment = Alignment(horizontal="center", vertical="center")
    ws["A3"].fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    ws["A3"].font = Font(color="FFFFFF", bold=True)

    start_col = 3
    for i in range(len(res)):
        end_col = start_col + 1
        start_col_letter = chr(64 + start_col)
        end_col_letter = chr(64 + end_col)
        ws.merge_cells(f"{start_col_letter}2:{end_col_letter}2")
        ws[f"{start_col_letter}2"] = res[i]
        ws[f"{start_col_letter}2"].alignment = Alignment(horizontal="center", vertical="center")
        start_col += 2

    row_labels = res
    row_index = 4
    for label in row_labels:
        ws.merge_cells(f"A{row_index}:A{row_index + 1}")
        ws[f"A{row_index}"] = label
        ws[f"A{row_index}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"B{row_index}"] = "P45"
        ws[f"B{row_index + 1}"] = "M45"
        row_index += 2

    wb.save(filename)
    print(f"Excel file '{filename}' created successfully!")


# Function to extract M3 values from a text file
def extract_m3_values(file_path):
    m3_values = {}
    current_trace = None

    with open(file_path, 'r') as file:
        for line in file:
            if 'Trc' in line:
                current_trace = line.split()[1]

            if 'M3' in line:
                parts = line.split()
                m3_values[current_trace] = float(parts[-2])
    print(f"M3 values extracted: {m3_values}")
    return m3_values


# Function to get the position of a trace from the mapping
def get_excel_position(trace, mapping):
    return mapping.get(trace, None)


# Function to update the Excel file with M3 values
def update_excel_with_values(excel_file, m3_values, filename_pattern, mapping):
    wb = load_workbook(excel_file)
    ws = wb.active

    row_labels = re.findall(r"(R\d+|Y\d+)", filename_pattern)

    if len(row_labels) < 2:
        print(f"Invalid filename pattern for {filename_pattern}")
        return

    target_row_label, target_col_label = row_labels

    # Debugging row and column indices
    row_indices = {}
    for row in range(4, ws.max_row + 1, 2):
        row_value = ws.cell(row=row, column=1).value
        if row_value == target_row_label:
            row_indices[target_row_label] = row

    column_indices = {}
    for col in range(3, ws.max_column + 1, 2):
        header_value = ws.cell(row=2, column=col).value
        if header_value == target_col_label:
            column_indices[target_col_label] = col
    print(f"Column indices: {column_indices}")

    p = 0

    for trace, m3_value in m3_values.items():
        print(f"Processing trace: {trace} with m3_value: {m3_value}")
        position = get_excel_position(trace, mapping)

        if not position:
            default_position = [("P45", "P45"), ("P45", "M45"), ("M45", "P45"), ("M45", "M45")]
            print(f"Position not found for trace: {trace}. Adding default positions.")
            # Add multiple default variations to mapping
            mapping[trace] = default_position[p]
            position = mapping[trace]
            p = p + 1
            print(f"Position for {trace}: {position}")
            print(f"Updated mapping: {mapping}")
        else:
            print(f"Position found for {trace}: {position}")

        sub_row_label, sub_col_label = position

        if target_row_label in row_indices and target_col_label in column_indices:
            row = row_indices[target_row_label] + (0 if sub_row_label == "P45" else 1)
            col = column_indices[target_col_label] + (0 if sub_col_label == "P45" else 1)
            ws.cell(row=row, column=col, value=m3_value)
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

    wb.save(excel_file)
    print(f"Updated Excel file '{excel_file}' with {filename_pattern} values!")


def get_interband(filename, fold):
    print(f"filepath from vna.py {filename}")
    print(f"folder directory from vna.py {fold}")
    base_name = os.path.basename(filename)  # e.g., 'R1_constant_R2.txt'
    # port_name = base_name.split('_')[1]  # e.g., 'R1'


    # Initial mapping dictionary
    mapping = {}

    excel_file_path = os.path.join(fold, f"\\interband.xlsx")
    if os.path.exists(excel_file_path):
        print("excel already exist")
    else:
        create_excel_sheet(excel_file_path,"text.txt")

    # for filename_pattern, text_file_path in text_files:
    time.sleep(10)
    m3_values = extract_m3_values(filename)
    update_excel_with_values(excel_file_path, m3_values, base_name,mapping)