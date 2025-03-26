from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os
import datetime
import sys

current_timestamp = datetime.datetime.now()
timestamp_str = current_timestamp.strftime("%Y%m%d_%H%M%S")

thick = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)
no_border = Border(
    top=Side(border_style=None),
    bottom=Side(border_style=None)
)

thick_border = Side(style='thick')
webdings_font = Font(name='Webdings')
symbol_font = Font(name='Symbol')


def process_row(row):
    """
    """
    row = list(row)
    if row[0] == 'S11':
        row[0] = "ISO"
        row[1:1] = [""]
        row[3:3] = ["dB", "", ""]
        row[7:7] = ["dB", "", ""]
        row[11:11] = ["dB", ""]
    elif row[0] == 'S12':
        row[0] = "VSWR"
        row[1:1] = ["Å"]
        row[3:3] = ["y", "Å"]
        row[6:6] = ["y", "Å"]
        row[9:9] = ["y"]
    elif row[0] == 'S22':
        row = row[1:]
    return row[:-1]


def get_worst_case_scenario(row):
    """
    """
    row = list(row)
    if row[0] == 'S11':
        row = row[-2:-1]
        row.insert(0, "ISO")
        row[1:1] = [""]
        row[3:3] = ["dB", ""]
    elif row[0] == 'S12':
        row = row[-2:-1]
        row.insert(0, "VSWR")
        row[1:1] = ["Å"]
        row[3:3] = ["y"]
    elif row[0] == 'S22':
        row = row[-1:]
    return row


def read_excel_files():
    """
    """
    list_of_items = []
    list_of_description = []
    for subdir, dir, list_files in os.walk(os.getcwd()):
        for files in list_files:
            if files.endswith(".xlsx") and not files.startswith("VSWR_Table"):
                components = subdir.split(os.sep)
                if components[-1] != "Interband":
                    wb = load_workbook(os.path.join(subdir, files), read_only=True)
                    ws = wb['Sheet1']
                    list_of_description.append(components[-3] + " " + components[-2] + " " + components[-1])
                    for index, row in enumerate(ws.iter_rows(values_only=True)):
                        if index == 0:
                            continue  # Skip the first row
                        list_of_items.append(list(row))
    if len(list_of_items) == 0:
        print("File not found")
        sys.exit(0)
    return list_of_items, list_of_description


def group_items(list_of_items):
    """
    """
    grouped_items = {}
    for each in list_of_items:
        if grouped_items and each[0] in grouped_items.keys():
            grouped_items[each[0]].append(each[1:])
        else:
            grouped_items[each[0]] = [each[1:]]
    return grouped_items


def sort_grouped_item(grouped_item, function_name):
    """
    """
    max_length = 0
    dup_grouped_item = grouped_item.copy()
    for key, items in dup_grouped_item.items():
        temp_var = []
        for list_of in items:
            temp_var.append(function_name(list_of))
            if len(items) > max_length:
                max_length = len(items)
        dup_grouped_item[key] = temp_var
    return dup_grouped_item, max_length / 3


def update_data(description, grouped_data, max_lenght=0):
    """
    """
    description_line = ["", "", "ALL", ""]
    for each_description in description:
        description_line.extend([each_description, "", "", "", "", "", "", "", "", "", "", ""])
    max_min_list = ["min", "", "", "", "mid", "", "", "", "max", "", "", ""]
    repeated_list = [item for _ in range(max_lenght) for item in max_min_list]
    min_list = ["", "", "", ""]
    min_list.extend(repeated_list)
    excel_data = [
        description_line,
        min_list
    ]
    for key, items in grouped_data.items():
        iso_path = ["", 'ISO', "", ""]
        vswr_path = [key, "VSWR", "+", "-"]
        for count, data in enumerate(items):
            if len(items) - 1 == count:
                break
            if data[0] == 'ISO':
                iso_path.extend(data[1:])
            elif data[0] == 'VSWR':
                data.insert(4, items[count + 1][0])
                data.insert(8, items[count + 1][1])
                data.append(items[count + 1][2])
                vswr_path.extend(data[1:])

        excel_data.append(iso_path)
        excel_data.append(vswr_path)
        date_item = ["DD/MM/YYYY", "", "", ""]
        date = ["", "Date", "", ""]
        repeated_list = ([item for _ in range(max_lenght * 3) for item in date_item])
        excel_data.append(["", "Version"])
        date.extend(repeated_list)
        excel_data.append(date)
    return excel_data


def update_worst_data(description, grouped_data, max_lenght=0):
    """
    """
    description_line = ["", "", "ALL", ""]
    for each_description in description:
        description_line.extend([each_description, "", "", ""])
    excel_data = [
        description_line,
    ]
    for key, items in grouped_data.items():
        iso_path = ["", 'ISO', "", ""]
        vswr_path = [key, "VSWR", "+", "-"]
        for count, data in enumerate(items):
            if len(items) - 1 == count:
                break
            if data[0] == 'ISO':
                iso_path.extend(data[1:])
            elif data[0] == 'VSWR':
                data.append(items[count + 1][0])
                vswr_path.extend(data[1:])

        excel_data.append(iso_path)
        excel_data.append(vswr_path)
        date_item = ["DD/MM/YYYY", "", "", ""]
        date = ["", "Date", "", ""]
        repeated_list = ([item for _ in range(max_lenght) for item in date_item])
        excel_data.append(["", "Version"])
        date.extend(repeated_list)
        excel_data.append(date)
    return excel_data


def all_cell_border(ws):
    """
    """
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thick


def excel_fill_data(excel_data, ws):
    """
    """
    ws.sheet_view.showGridLines = False
    fill_color = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    for row_num, row_data in enumerate(excel_data, 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
            if cell_value == 'y':
                cell.font = webdings_font
            elif cell_value == "Å":
                cell.font = symbol_font
            # ws.cell(row=row_num, column=col_num).fill = fill_color
    return ws


def apply_thick_border(ws, start_row, start_col, end_row, end_col):
    # Apply top border
    # start_row, start_col = 1, 1
    # end_row, end_col = 3, 3
    # Define the border
    thick = Side(style='thick')

    # Apply borders only to the outer edges of the range
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)

            # Apply borders only to the edges
            border = Border(
                top=thick if row == start_row else None,
                bottom=thick if row == end_row else None,
                left=thick if col == start_col else None,
                right=thick if col == end_col else None
            )

            cell.border = border  # Assign the border to the cell


def auto_adjust_column_width(ws):
    """Auto-adjust column width while setting columns with 'Å' to 50 pixels."""
    column_widths = {}
    special_columns = set()  # Track columns that contain 'Å'
    # Get a set of merged cell coordinates to ignore them
    merged_cells = set()
    for merged_range in ws.merged_cells.ranges:
        for row in ws.iter_rows(min_row=merged_range.min_row, max_row=merged_range.max_row,
                                min_col=merged_range.min_col, max_col=merged_range.max_col):
            for cell in row:
                merged_cells.add(cell.coordinate)  # Store merged cell coordinates

    # Determine max length for each column (excluding merged cells)


    for row in ws.iter_rows():
        for cell in row:
            if cell.value and cell.coordinate not in merged_cells:
                col_letter = get_column_letter(cell.column)
                # adjusted_length = get_adjusted_length(cell.value)
                column_widths[col_letter] = max(column_widths.get(col_letter, 0), len(str(cell.value)))

                # Mark columns that contain 'Å'
                if "Å" in str(cell.value):
                    special_columns.add(col_letter)

    # Apply calculated column widths
    for col_letter in ws.column_dimensions:
        if col_letter in special_columns:
            ws.column_dimensions[col_letter].width = (50 / 7) + 2 # Convert pixels to Excel width
        else:
            ws.column_dimensions[col_letter].width = column_widths.get(col_letter, 10) + 2  # Auto width with padding

    # Apply the calculated column widths
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width + 2  # Adding padding


def add_border(ws, start_cell, end_cell):
    start_col, start_row = start_cell[0], int(start_cell[1])
    end_col, end_row = end_cell[0], int(end_cell[1])
    thick = Side(style='thick')
    # Top border
    for col in range(ord(start_col), ord(end_col) + 1):
        ws[f"{chr(col)}{start_row}"].border = Border(top=thick)

    # Bottom border
    for col in range(ord(start_col), ord(end_col) + 1):
        ws[f"{chr(col)}{end_row}"].border = Border(bottom=thick)
    # Left border
    for row in range(start_row, end_row + 1):
        ws[f"{start_col}{row}"].border = Border(left=thick)

    # Right border
    for row in range(start_row, end_row + 1):
        ws[f"{end_col}{row}"].border = Border(right=thick)


if __name__ == "__main__":
    main_folder = r"C:\Users\harini.puche\Documents"
    user_name = input("Enter the user name: ")
    print(os.listdir(os.path.join(main_folder, user_name)))
    antenna_model_number = input("Antenna model number: ")
    os.chdir(os.path.join(main_folder, user_name, antenna_model_number))
    group_item, description = read_excel_files()

    complete_group = group_items(group_item)
    sorted_dict, length = sort_grouped_item(complete_group, process_row)
    excel_data = update_data(description, sorted_dict, int(length))
    wb = Workbook()
    ws = wb.active
    ws.title = "VSWR & ISOLATION"
    # ws.sheet_view.showGridLines = False
    ws = excel_fill_data(excel_data, ws)
    last_ws = ws.max_row
    last_col = ws.max_column
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
            # Change "DD/MM/YYYY" to 'isdate'
            if cell.value in ["min", "mid", "max", "DD/MM/YYYY"]:
                ws.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                               end_column=cell.column + 3)
                ws.alignment = Alignment(horizontal='center', vertical='center')
                if cell.row < last_ws and cell.column < last_col:
                    apply_thick_border(ws, cell.row + 1, cell.column, cell.row + 4, cell.column + 4)
            if cell.value in ["dB"]:
                next_cell = ws.cell(row=cell.row, column=cell.column - 1)
                if float(next_cell.value) > 25.9:
                    next_cell.font = Font(color="00B050")  # Green
                elif float(next_cell.value) < 25:
                    next_cell.font = Font(color="FF0000")  # Red
                elif float(next_cell.value) > 25 and int(next_cell.value) < 25.9:
                    next_cell.font = Font(color="FFFF00")  # Yellow
            if cell.row == 1 and cell.value not in ["ALL", "", None]:
                ws.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                               end_column=cell.column + 11)
                ws.alignment = Alignment(horizontal='center', vertical='center')
            elif cell.value in ["ALL"]:
                ws.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                               end_column=cell.column + 1)
                ws.alignment = Alignment(horizontal='center', vertical='center')
                # if cell.column == 1:
            #     if cell.value != None and len(cell.value) > 0:
            #         ws.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row + 3, end_column=cell.column)
            #         ws.alignment = Alignment(horizontal='center', vertical='center')
            if cell.value == 'Å' or cell.value == 'y':
                next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                if float(next_cell.value) < 1.47:
                    next_cell.font = Font(color="00B050")
                elif float(next_cell.value) >= 1.5:
                    next_cell.font = Font(color="FF0000")
                elif float(next_cell.value) > 1.47 and int(next_cell.value) < 1.49:
                    next_cell.font = Font(color="FFFF00")
    fill_color = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    ws.delete_cols(last_col)
    # all_cell_border(ws)
    # ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column = int(length)*12+4)
    for row in range(3, last_ws + 1, 4):
        apply_thick_border(ws, row, 1, row + 3, 1)
        apply_thick_border(ws, row, 2, row + 3, 2)
        apply_thick_border(ws, row, 3, row + 3, 4)
    for row in range(1, 3):
        for col in range(5, int(length) * 12 + 1 + 4):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill_color
            cell.border = thick
    apply_thick_border(ws, 1, 1, 2, 4)

    sorted_dict, length = sort_grouped_item(complete_group, get_worst_case_scenario)
    sheet2 = update_worst_data(description, sorted_dict, int(length))
    # ws = wb.create_sheet(title="VSWR and Isolation")
    ws2 = wb.create_sheet(title="VSWR and Isolation(Worst case)")
    ws2 = excel_fill_data(sheet2, ws2)
    # Save the workbook
    last_row_num = ws2.max_row
    last_col_num = ws2.max_column
    for row in ws2.iter_rows():

        for cell in row:
            cell.alignment = Alignment(horizontal='center')
            # Change "DD/MM/YYYY" to 'isdate'
            if cell.value in ["DD/MM/YYYY"]:
                ws2.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                                end_column=cell.column + 3)
                ws2.alignment = Alignment(horizontal='center', vertical='center')
                if cell.row < last_row_num:
                    apply_thick_border(ws2, cell.row + 1, cell.column, cell.row + 4, cell.column + 4)
            if cell.value in description:
                apply_thick_border(ws2, cell.row + 1, cell.column, cell.row + 4, cell.column + 4)
            if cell.value in ["dB"]:
                next_cell = ws2.cell(row=cell.row, column=cell.column - 1)
                if float(next_cell.value) > 25.9:
                    next_cell.font = Font(color="00B050")  # Green
                elif float(next_cell.value) < 25:
                    next_cell.font = Font(color="FF0000")  # Red
                elif float(next_cell.value) > 25 and int(next_cell.value) < 25.9:
                    next_cell.font = Font(color="FFFF00")  # Yellow
            if cell.row == 1 and cell.value not in ["ALL", "", None]:
                ws2.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                                end_column=cell.column + 3)
                ws2.alignment = Alignment(horizontal='center', vertical='center')
            elif cell.value in ["ALL"]:
                ws2.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row,
                                end_column=cell.column + 1)
                ws2.alignment = Alignment(horizontal='center', vertical='center')
                # if cell.column == 1:
            #     if cell.value != None and len(cell.value) > 0:
            #         ws2.merge_cells(start_row=cell.row, start_column=cell.column, end_row=cell.row + 3, end_column=cell.column)
            #         ws2.alignment = Alignment(horizontal='center', vertical='center')
            if cell.value == 'Å' or cell.value == 'y':
                next_cell = ws2.cell(row=cell.row, column=cell.column + 1)
                if float(next_cell.value) < 1.47:
                    next_cell.font = Font(color="00B050")
                elif float(next_cell.value) >= 1.5:
                    next_cell.font = Font(color="FF0000")
                elif float(next_cell.value) > 1.47 and int(next_cell.value) < 1.49:
                    next_cell.font = Font(color="FFFF00")
    fill_color = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    ws2.delete_cols(last_col_num)
    add_border(ws2, "A1", "D1")
    for row in range(2, last_row_num + 1, 4):
        apply_thick_border(ws2, row, 1, row + 3, 1)
        apply_thick_border(ws2, row, 2, row + 3, 2)
        apply_thick_border(ws2, row, 3, row + 3, 4)
    # all_cell_border(ws)
    # ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column = int(length)*12+4)
    for row in range(1, 2):
        for col in range(5, int(length) * 4 + 4):
            cell = ws2.cell(row=row, column=col)
            cell.fill = fill_color
            cell.border = thick

    auto_adjust_column_width(ws)
    auto_adjust_column_width(ws2)
    wb.save("VSWR_Table_" + str(timestamp_str) + ".xlsx")